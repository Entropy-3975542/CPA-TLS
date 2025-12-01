import sys
import subprocess
import re
import time

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import LineChart, Reference
import scipy.stats as stats
import statistics

def run_openssl_s_time_complete(port, seconds=3):
    """
    Run openssl s_time command and capture all output
    """
    tmp_file = "tmp_runtime_HS.txt"
    cmd = f"gnome-terminal --wait -- bash -c 'openssl s_time -connect localhost:{port} -new -time {seconds} | tee {tmp_file}; exit'"
    
    try:
        # Launch new terminal
        process = subprocess.Popen(cmd, shell=True)
        print("Testing in progress, please wait...", end="", flush=True)
        process.wait()
        
        full_output = ""
        with open(f'{tmp_file}', 'r') as file:
            full_output = file.read()
        
        '''
        # Print all output for debugging
        print("\r=== Full Output ===")
        print(full_output)
        '''
            
        # Parse complete output
        parsed_data = parse_openssl_output(full_output)
        
        # Delete temporary file
        subprocess.run(f"rm {tmp_file}", shell=True)
        
        return parsed_data
        
    except subprocess.CalledProcessError as e:
        print(f"Command execution failed: {e}")
        return None

def parse_openssl_output(full_output):
    """
    Parse openssl speed output
    """
    data = {}
    
    # Match operation count and time in real-time progress lines
    # Format: "1995 connections in 0.29s; 6879.31 connections/user sec, bytes read 0"
    connections_pattern = r'(\d+).*?in\s*(\d+\.\d+)s'
    
    # Extract real-time progress data
    connections_match = re.search(connections_pattern, full_output)
    if connections_match:
        data['connections'] = int(connections_match.group(1))
        data['time'] = float(connections_match.group(2))
        data['us_per_connections'] = data['time']/data['connections']*1e6
        print(f"\rCollecting connection statistics:\n{data['connections']} connections in {data['time']}s, {data['us_per_connections']} μs/connection")
    
    return data

def get_runtime(scheme, seconds=3, runs=5):
    """Run command multiple times and collect results"""
    # Generate temporary certificates
    subprocess.run("openssl ecparam -name prime256v1 -genkey -noout -out tmp_server.key", shell=True)
    subprocess.run("openssl req -new -x509 -sha256 -key tmp_server.key -out tmp_server.crt -days 365 -subj '/CN=localhost'", shell=True)
    
    ports = [4443, 4444, 4445, 4446, 4447]
    process = None
    selected_port = None
    
    for port in ports:
        # Start server
        cmd = f"openssl s_server -cert tmp_server.crt -key tmp_server.key -accept {port} -www -groups {scheme}"
        try:
            # Launch new terminal
            process = subprocess.Popen(cmd, shell=True)
            # Wait a moment to see if the server stays up
            time.sleep(1)
            # Check if the process is still running
            if process.poll() is None:
                selected_port = port
                print(f"Server started successfully on port {selected_port}")
                break
            else:
                # Server process has exited, try next port
                print(f"Port {port} is busy, trying next...")
                time.sleep(1)
                process = None
        except Exception as e:
            print(f"Error starting server on port {port}: {e}")
            process = None
            
    if process is None:
        print("Failed to start server on any port")
        return None
    
    # Begin testing with selected_port
    all_results = []
    
    start_time = time.time()
    for i in range(runs):
        print(f"\nRun {i+1}...")
        result = run_openssl_s_time_complete(selected_port, seconds)
        if result:
            all_results.append(result)
            
        progress = (i + 1)/runs
        elapsed = time.time() - start_time
        eta = elapsed * (runs - i - 1) / (i + 1)
        
        print(f"Progress: {progress*100:.1f}% | {i+1}/{runs} | Elapsed: {format_time(elapsed)} | ETA: {format_time(eta)}")
    
    # Stop server and delete temporary certificates
    process.kill()
    subprocess.run(f"fuser -k {selected_port}/tcp", shell=True)
    subprocess.run("rm tmp_server.key tmp_server.crt", shell=True)
    
    return all_results
    
def format_time(seconds):
    """
    Format seconds into a human-readable time string
    Rules:
    - Less than 60 seconds: display seconds
    - 60 seconds to 60 minutes: display minutes and seconds
    - More than 60 minutes: display hours and minutes
    """
    if seconds < 60:
        return f"{seconds:.1f}s"
    elif seconds < 3600:  # 60 minutes * 60 seconds = 3600 seconds
        minutes = int(seconds // 60)
        remaining_seconds = seconds % 60
        return f"{minutes}m {remaining_seconds:.1f}s"
    else:
        hours = int(seconds // 3600)
        minutes = int((seconds % 3600) // 60)
        remaining_seconds = (seconds % 3600) % 60
        return f"{hours}h {minutes}m {remaining_seconds:.1f}s"

def save_to_excel_openpyxl(scheme, data):
    """
    Create detailed Excel report using openpyxl
    """
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"{scheme} Handshake Performance Test Results"
    
    # Set headers
    headers = ['Test Run', 'Time (µs/connection)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Fill data
    times = []
    
    for i, item in enumerate(data, 2):  # Start from row 2
        _time = item['us_per_connections']
        
        ws.cell(row=i, column=1, value=f"Run {i-1}")
        ws.cell(row=i, column=2, value=_time)
        
        times.append(_time)
    
    # Add statistical information
    stat_row = len(data) + 3
    
    res = calculate_CI(times)
    print(f"CV: {res[1]/res[0]*100:.1f}%")
    print(f"Mean: {res[0]:.2f}, Std: {res[1]:.2f}, 95% CI: [{res[2]:.2f}, {res[3]:.2f}], ME: {res[4]:.2f}")
    _stats = [
        ("Statistic", "Time (µs/connection)"),
        ("Mean", res[0]),
        ("Standard Deviation", res[1]),
        ("95% CI Lower Bound", res[2]),
        ("95% CI Upper Bound", res[3]),
        ("Margin of Error", res[4])
    ]
    
    for row_offset, stat_row_data in enumerate(_stats):
        for col, value in enumerate(stat_row_data, 1):
            cell = ws.cell(row=stat_row + row_offset, column=col, value=value)
            if row_offset == 0:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    # Set column width
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    
    # Save file
    filename = f"{scheme}_HS_performance_data.xlsx"
    wb.save(filename)
    print(f"Detailed report saved to {filename}")
    
    return {
        'times': times
    }
    
def calculate_CI(data, confidence_level=0.95):
    """
    Calculate confidence interval (using t-distribution)
    """
    n = len(data)
    if n < 2:
        return (0, 0)  # Insufficient samples
    
    mean = statistics.mean(data)
    stdev = statistics.stdev(data)
    
    # Use t-distribution critical value
    t_critical = stats.t.ppf((1 + confidence_level) / 2, df=n-1)
    
    margin_of_error = t_critical * (stdev / (n ** 0.5))
    
    ci_lower = mean - margin_of_error
    ci_upper = mean + margin_of_error
    
    return mean, stdev, ci_lower, ci_upper, margin_of_error

# Run test
if __name__ == "__main__":
    schemes = {"0": "X25519MLKEM768", "1": "SecP256r1MLKEM768", "2": "SecP384r1MLKEM1024"}
    scheme = sys.argv[1] if len(sys.argv) > 1 else ""
    while scheme not in schemes.values():
        tmp = input(f"Please choose from {schemes}:\n")
        if tmp in ["0", "1", "2"]:
            scheme = schemes[tmp]
        else:
            scheme = tmp
        
    seconds = 3
    runs = 50
    print(f"Test started! scheme = {scheme} -time = {seconds}, runs = {runs}")
    print("="*20)
    results = get_runtime(scheme, seconds, runs)
    
    print("Writing results to file")
    save_to_excel_openpyxl(scheme, results)
    print("Done")
