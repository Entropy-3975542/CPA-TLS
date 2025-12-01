import sys
import subprocess
import re
import time

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import LineChart, Reference
import scipy.stats as stats
import statistics

def run_openssl_speed_complete(scheme="X25519MLKEM768", seconds=3):
    """
    Run openssl speed command and capture all output (including real-time progress)
    """
    cmd = ["openssl", "speed", "-seconds", f"{seconds}", f"{scheme}"]
    
    try:
        # Execute command and capture both stdout and stderr
        result = subprocess.run(
            cmd, 
            check=True, 
            stdout=subprocess.PIPE, 
            stderr=subprocess.PIPE,  # Explicitly capture stderr
            text=True, 
            timeout=60
        )
        
        '''
        # Print all output for debugging
        print("=== Standard Output (stdout) ===")
        print(result.stdout)
        print("\n=== Standard Error (stderr) ===")
        print(result.stderr)
        print("\n" + "="*50 + "\n")
        '''
        
        # Combine all output for parsing
        full_output = result.stdout + "\n" + result.stderr
        
        # Parse complete output
        parsed_data = parse_openssl_output(full_output)
        
        return parsed_data
        
    except subprocess.CalledProcessError as e:
        print(f"Command execution failed: {e}")
        return None

def parse_openssl_output(full_output):
    """
    Parse openssl speed output
    """
    data = {}
    
    # Match operation count and time in real-time progress lines
    # Format: "Doing X25519MLKEM768 encaps ops for 3s: 22061 X25519MLKEM768 KEM encaps ops in 2.96s"
    encaps_progress_pattern = r'Doing.*encaps.*?: (\d+).*?in\s*(\d+\.\d+)s'
    decaps_progress_pattern = r'Doing.*decaps.*?: (\d+).*?in\s*(\d+\.\d+)s'
    
    # Extract real-time progress data
    encaps_match = re.search(encaps_progress_pattern, full_output)
    if encaps_match:
        data['encaps_ops'] = int(encaps_match.group(1))
        data['encaps_time'] = float(encaps_match.group(2))
        data['encaps_us_per_ops'] = data['encaps_time']/data['encaps_ops']*1e6
        print(f"Found encapsulation progress: {data['encaps_ops']} operations, time {data['encaps_time']} seconds, {data['encaps_us_per_ops']} μs/ops")
    
    decaps_match = re.search(decaps_progress_pattern, full_output)
    if decaps_match:
        data['decaps_ops'] = int(decaps_match.group(1))
        data['decaps_time'] = float(decaps_match.group(2))
        data['decaps_us_per_ops'] = data['decaps_time']/data['decaps_ops']*1e6
        print(f"Found decapsulation progress: {data['decaps_ops']} operations, time {data['decaps_time']} seconds, {data['decaps_us_per_ops']} μs/ops")
    
    return data

def get_runtime(scheme="X25519MLKEM768", seconds=3, runs=5):
    """Run command multiple times and collect results"""
    all_results = []
    
    start_time = time.time()
    for i in range(runs):
        print(f"\nRun {i+1}...")
        result = run_openssl_speed_complete(scheme, seconds)
        if result:
            all_results.append(result)
            
        progress = (i + 1)/runs
        elapsed = time.time() - start_time
        eta = elapsed * (runs - i - 1) / (i + 1)
        
        print(f"Progress: {progress*100:.1f}% | {i+1}/{runs} | Elapsed: {format_time(elapsed)} | ETA: {format_time(eta)}")
    
    return all_results
    
def format_time(seconds):
    """
    Format seconds into a human-readable time string
    Rules:
    - Less than 60 seconds: display seconds
    - 60 seconds to 60 minutes: display minutes and seconds
    - More than 60 minutes: display hours and minutes
    """
    if seconds < 60:
        return f"{seconds:.1f}s"
    elif seconds < 3600:  # 60 minutes * 60 seconds = 3600 seconds
        minutes = int(seconds // 60)
        remaining_seconds = seconds % 60
        return f"{minutes}m {remaining_seconds:.1f}s"
    else:
        hours = int(seconds // 3600)
        minutes = int((seconds % 3600) // 60)
        remaining_seconds = (seconds % 3600) % 60
        return f"{hours}h {minutes}m {remaining_seconds:.1f}s"

def save_to_excel_openpyxl(data, scheme):
    """
    Create detailed Excel report using openpyxl
    """
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"{scheme} Performance Test Results"
    
    # Set headers
    headers = ['Test Run', 'Encapsulation Time (µs)', 'Decapsulation Time (µs)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Fill data
    encaps_times = []
    decaps_times = []
    
    for i, item in enumerate(data, 2):  # Start from row 2
        encaps_time = item['encaps_us_per_ops']
        decaps_time = item['decaps_us_per_ops']
        
        ws.cell(row=i, column=1, value=f"Run {i-1}")
        ws.cell(row=i, column=2, value=encaps_time)
        ws.cell(row=i, column=3, value=decaps_time)
        
        encaps_times.append(encaps_time)
        decaps_times.append(decaps_time)
    
    # Add statistical information
    stat_row = len(data) + 3
    
    res1 = calculate_CI(encaps_times)
    print(f"Encaps (CV: {res1[1]/res1[0]*100:.1f}%):")
    print(f"Mean: {res1[0]:.2f}, Std: {res1[1]:.2f}, 95% CI: [{res1[2]:.2f}, {res1[3]:.2f}], ME: {res1[4]:.2f}")
    res2 = calculate_CI(decaps_times)
    print(f"Decaps (CV: {res2[1]/res2[0]*100:.1f}%):")
    print(f"Mean: {res2[0]:.2f}, Std: {res2[1]:.2f}, 95% CI: [{res2[2]:.2f}, {res2[3]:.2f}], ME: {res2[4]:.2f}")
    _stats = [
        ("Statistic", "Encapsulation Time", "Decapsulation Time"),
        ("Mean", res1[0], res2[0]),
        ("Standard Deviation", res1[1], res2[1]),
        ("95% CI Lower Bound", res1[2], res2[2]),
        ("95% CI Upper Bound", res1[3], res2[3]),
        ("Margin of Error", res1[4], res2[4])
    ]
    
    for row_offset, stat_row_data in enumerate(_stats):
        for col, value in enumerate(stat_row_data, 1):
            cell = ws.cell(row=stat_row + row_offset, column=col, value=value)
            if row_offset == 0:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    # Set column width
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    
    # Save file
    filename = f"{scheme}_performance_data.xlsx"
    wb.save(filename)
    print(f"Detailed report saved to {filename}")
    
    return {
        'encaps_times': encaps_times,
        'decaps_times': decaps_times
    }
    
def calculate_CI(data, confidence_level=0.95):
    """
    Calculate confidence interval (using t-distribution)
    """
    n = len(data)
    if n < 2:
        return (0, 0)  # Insufficient samples
    
    mean = statistics.mean(data)
    stdev = statistics.stdev(data)
    
    # Use t-distribution critical value
    t_critical = stats.t.ppf((1 + confidence_level) / 2, df=n-1)
    
    margin_of_error = t_critical * (stdev / (n ** 0.5))
    
    ci_lower = mean - margin_of_error
    ci_upper = mean + margin_of_error
    
    return mean, stdev, ci_lower, ci_upper, margin_of_error

# Run test
if __name__ == "__main__":
    schemes = {"0": "X25519MLKEM768", "1": "SecP256r1MLKEM768", "2": "SecP384r1MLKEM1024"}
    scheme = sys.argv[1] if len(sys.argv) > 1 else ""
    while scheme not in schemes.values():
        tmp = input(f"Please choose from {schemes}:\n")
        if tmp in ["0", "1", "2"]:
            scheme = schemes[tmp]
        else:
            scheme = tmp
            
    seconds = 3
    runs = 50
    print(f"Test started! scheme = {scheme}, seconds = {seconds}, runs = {runs}")
    print("="*20)
    results = get_runtime(scheme, seconds, runs)
    
    print("Writing results to file")
    save_to_excel_openpyxl(results, scheme)
    print("Done")
